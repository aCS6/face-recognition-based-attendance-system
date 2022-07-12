import face_recognition
from PIL import Image, ImageDraw
import os
import sys
from pathlib import Path
from datetime import datetime
from datetime import date
from openpyxl import load_workbook
import shutil
import subprocess



recognizingImagePath = sys.argv[1]





#get the directories
path = '/home/amir/Documents/face/img/known' # using pwd command in linux

files = []  # CONTAINS FILE NAME WITH DIRECTORY
# r=root, d = directories, f = files
for r, d, f in os.walk(path):
    for file in f:
        if '.jpg' in file:
            files.append(os.path.join(r, file))
        elif '.jpeg' in file:
            files.append(os.path.join(r, file))
        elif '.png' in file:
            files.append(os.path.join(r, file))



entries = []
image = []
encoding = []
i=0

allStudent=[]
#get file names without extension
for f in files:
    entries.append(Path(f).stem)
for f in files:
    # image.append(face_recognition.load_image_file(f))
    encoding.append(face_recognition.face_encodings(face_recognition.load_image_file(f))[0])
    i = i+1

known_face_encodings = encoding
known_face_names = entries
test_image = face_recognition.load_image_file(recognizingImagePath)

#find faces in the test image
face_locations = face_recognition.face_locations(test_image)
face_encodings = face_recognition.face_encodings(test_image,face_locations)

# covert to PIL format
pil_image = Image.fromarray(test_image)

#create a ImageDraw instance
draw = ImageDraw.Draw(pil_image)

# Loop through faces in test image
for(top,right,bottom,left),face_encoding in zip(face_locations,face_encodings):
    matches = face_recognition.compare_faces(known_face_encodings,face_encoding)

    name = "Unknown"

    if True in matches:  #if faces matches
        first_match_index = matches.index(True) 
        name = known_face_names[first_match_index]

        allStudent.append(name)
    

    #draw Box
    draw.rectangle(((left,top),(right,bottom)),outline=(0,0,1))

    #draw label
    text_width,text_height = draw.textsize(name)
    draw.rectangle(((left, bottom - text_height - 10),(right,bottom)),fill=(0,0,1),outline=(0,0,1))
    draw.text((left+6,bottom - text_height - 5),name,fill=(255,255,255,255))

del draw


allStudent = list(dict.fromkeys(allStudent)) # remove duplicates
allStudent = list(map(int, allStudent)) # convert all elements from string to int
allStudent.sort() # sorting the elements


now = datetime.now()
# Display image
pil_image.save(f'{now}.jpg')


today = date.today()
# dd/mm/yy
getDate = today.strftime("%d/%m/%y")

#get total class
with open('/home/amir/Documents/face/doc/total class.txt','r') as f:
    f_content = f.read()
f_content = int(f_content)

#get excel sheet location
filepath = "/home/amir/Documents/face/doc/python.xlsx"

wb = load_workbook(filepath)
sheet = wb.active

row = 1
col = f_content + 1 #indicating class number getting from txt file

dateData = sheet.cell(row=row,column=col)
dateData.value = getDate
row = row + 1

for i in allStudent:
    data = sheet.cell(row=row,column=col)
    data.value = i
    row = row + 1

wb.save(filepath)


f_content = str(col)

with open('/home/amir/Documents/face/doc/total class.txt','w') as f:
    f.write(f_content)

source_path = '/home/amir/Documents/face/forms'
destination_path = '/home/amir/Documents/face/img/taken' # using pwd command in linux

files_for_moving = []  # CONTAINS FILE NAME WITH DIRECTORY
# r=root, d = directories, f = files
for r, d, f in os.walk(source_path):
    for file in f:
        if '.jpg' in file:
            files_for_moving.append(os.path.join(r, file))

for f in files_for_moving:
    shutil.move(f,destination_path)



f_content = getDate

with open('/home/amir/Documents/face/doc/lastdate.txt','w') as f:
    f.write(f_content)

try:
    subprocess.check_call(["pkill", "-f", "takeAtten.py"])  # Enter the script name to terminate. Here (p1.py)
except:
    pass

pil_image.show()